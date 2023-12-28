# ***************************************************************************
# *   Copyright (c) 2023 Paul Ebbers paul.ebbers@gmail.com                  *
# *                                                                         *
# *   This file is part of the FreeCAD CAx development system.              *
# *                                                                         *
# *   This program is free software; you can redistribute it and/or modify  *
# *   it under the terms of the GNU Lesser General Public License (LGPL)    *
# *   as published by the Free Software Foundation; either version 2 of     *
# *   the License, or (at your option) any later version.                   *
# *   for detail see the LICENCE text file.                                 *
# *                                                                         *
# *   FreeCAD is distributed in the hope that it will be useful,            *
# *   but WITHOUT ANY WARRANTY; without even the implied warranty of        *
# *   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the         *
# *   GNU Lesser General Public License for more details.                   *
# *                                                                         *
# *   You should have received a copy of the GNU Library General Public     *
# *   License along with FreeCAD; if not, write to the Free Software        *
# *   Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  *
# *   USA                                                                   *
# *                                                                         *
# ***************************************************************************/


def Mbox(text, title="", style=0, default="", stringList="[,]"):
    """
    Message Styles:\n
    0 : OK                          (text, title, style)\n
    1 : Yes | No                    (text, title, style)\n
    2 : Inputbox                    (text, title, style, default)\n
    3 : Inputbox with dropdown      (text, title, style, default, stringlist)\n
    """
    from PySide2.QtWidgets import QMessageBox, QInputDialog
    import FreeCAD as App

    Parent = App.Gui.getMainWindow
    
    if style == 0:
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(text )
        msgBox.setWindowTitle(title)
        
        reply  = msgBox.exec_()
        return reply
    if style == 1:
        msgBox = QMessageBox()
        msgBox.setIcon(QMessageBox.Information)
        msgBox.setText(text )
        msgBox.setWindowTitle(title)
        msgBox.setStandardButtons(QMessageBox.Yes | QMessageBox.No)        
        msgBox.setDefaultButton(QMessageBox.No)
        
        reply = msgBox.exec_()
        if reply == QMessageBox.Yes:
            return "yes"
        if reply == QMessageBox.No:
            return "no"
    if style == 2:
        reply = QInputDialog.getText(Parent, title, text, text=default)
        if reply[1]:
            # user clicked OK
            replyText = reply[0]
        else:
            # user clicked Cancel
            replyText = reply[0]  # which will be "" if they clicked Cancel
        return str(replyText)
    if style == 3:
        reply = QInputDialog.getItem(Parent, title, text, stringList, 1, True)
        if reply[1]:
            # user clicked OK
            replyText = reply[0]
        else:
            # user clicked Cancel
            replyText = reply[0]  # which will be "" if they clicked Cancel
        return str(replyText)


def SaveDialog(files, OverWrite: bool = True):
    """
    files must be like:\n
    files = [\n
        ('All Files', '*.*'),\n
        ('Python Files', '*.py'),\n
        ('Text Document', '*.txt')\n
    ]\n
    \n
    OverWrite:\n
    If True, file will be overwritten\n
    If False, only the path+filename will be returned\n
    """
    import tkinter as tk
    from tkinter.filedialog import asksaveasfile
    from tkinter.filedialog import askopenfilename

    # Create the window
    root = tk.Tk()
    # Hide the window
    root.withdraw()

    if OverWrite is True:
        file = asksaveasfile(filetypes=files, defaultextension=files)
        if file is not None:
            return file.name
    if OverWrite is False:
        file = askopenfilename(filetypes=files, defaultextension=files)
        if file is not None:
            return file


def GetLetterFromNumber(number: int, UCase: bool = True):
    from openpyxl.utils import get_column_letter

    Letter = get_column_letter(number)

    # If UCase is true, convert to upper case
    if UCase is True:
        Letter = Letter.upper()

    return Letter


def GetNumberFromLetter(Letter):
    from openpyxl.utils import column_index_from_string

    Number = column_index_from_string(Letter)

    return Number


def GetA1fromR1C1(input: str) -> str:
    if input[:1] == "'":
        input = input[1:]
    try:
        input = input.upper()
        ColumnPosition = input.find("C")
        RowNumber = int(input[1:(ColumnPosition)])
        ColumnNumber = int(input[(ColumnPosition + 1) :])

        ColumnLetter = GetLetterFromNumber(ColumnNumber)

        return str(ColumnLetter + str(RowNumber))
    except Exception:
        return ""


def CheckIfWorkbookExists(FullFileName: str, CreateIfNone: bool = True):
    import os
    from openpyxl import Workbook

    result = False
    try:
        result = os.path.exists(FullFileName)
    except Exception:
        if CreateIfNone is True:
            Filter = [
                ("Excel", "*.xlsx"),
                (
                    "Excel Macro-enabled Workbook",
                    "*.xlsm",
                ),
            ]
            FullFileName = SaveDialog(Filter)
            if FullFileName.strip():
                wb = Workbook(str(FullFileName))
                wb.save(FullFileName)
                wb.close()
                result = True
        if CreateIfNone is False:
            result = False
    return result


def ColorConvertor(ColorRGB: [], Alpha: float = 1) -> ():
    """
    A single function to convert colors to rgba colors as a tuple of float from 0-1
    ColorRGB:   [255,255,255]
    Alpha:      0-1
    """
    from matplotlib import colors as mcolors

    ColorRed = ColorRGB[0] / 255
    colorGreen = ColorRGB[1] / 255
    colorBlue = ColorRGB[2] / 255

    color = (ColorRed, colorGreen, colorBlue)

    result = mcolors.to_rgba(color, Alpha)

    return result


def OpenFile(FileName: str):
    """
    Filename = full path with filename as string
    """
    import subprocess
    import os
    import platform

    if os.path.exists(FileName):
        if platform.system() == "Darwin":  # macOS
            subprocess.call(("open", FileName))
        elif platform.system() == "Windows":  # Windows
            os.startfile(FileName)
        else:  # linux variants
            subprocess.call(("xdg-open", FileName))
    else:
        print(f"Error: {FileName} does not exist.")


def SetColumnWidth_SpreadSheet(sheet, cellValue: str, factor=10) -> bool:
    result = False
    if cellValue == "" or sheet is None:
        result = False
        return result

    # Calculate the text length needed.
    length = int(len(cellValue) * factor)

    # Set the column width
    sheet.setColumnWidth("A", length)

    # Recompute the sheet
    sheet.recompute()

    return True


def Print(Input: str, Type: str = ""):
    """_summary_

    Args:
        Input (str): Text to print.\n
        Type (str, optional): Type of message. (enter Warning, Error or Log). Defaults to "".
    """
    import FreeCAD as App

    if Type == "Warning":
        App.Console.PrintWarning(Input + "\n")
    elif Type == "Error":
        App.Console.PrintError(Input + "\n")
    elif Type == "Log":
        App.Console.PrintLog(Input + "\n")
    else:
        App.Console.PrintMessage(Input + "\n")
